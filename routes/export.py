from __future__ import annotations

from datetime import datetime
from io import BytesIO

from flask import Blueprint, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from firebase_config import ATTENDANCE, MARKS, STUDENTS, SUBJECTS, db

export_bp = Blueprint("export", __name__)


def _subject_name(subject_code: str) -> str:
    doc = db.collection(SUBJECTS).document(subject_code).get()
    if doc.exists:
        return (doc.to_dict() or {}).get("name", subject_code)
    return subject_code


def _grade_from_percentage(pct: float) -> str:
    if pct >= 90:
        return "O"
    if pct >= 80:
        return "A+"
    if pct >= 70:
        return "A"
    if pct >= 60:
        return "B+"
    if pct >= 55:
        return "B"
    if pct >= 50:
        return "C"
    if pct >= 40:
        return "P"
    return "F"


def _attendance_rows(subject_code: str) -> list[dict]:
    docs = list(db.collection(ATTENDANCE).where("subject_code", "==", subject_code).stream())
    class_dates = {doc.to_dict().get("date") for doc in docs if doc.to_dict().get("date")}
    total_classes = len(class_dates)
    present_by_usn: dict[str, int] = {}
    for doc in docs:
        data = doc.to_dict() or {}
        if data.get("status") == "present":
            usn = (data.get("usn") or "").upper()
            present_by_usn[usn] = present_by_usn.get(usn, 0) + 1

    rows = []
    for doc in db.collection(STUDENTS).order_by("usn").stream():
        data = doc.to_dict() or {}
        usn = (data.get("usn") or doc.id).upper()
        attended = present_by_usn.get(usn, 0)
        pct = round((attended / total_classes) * 100, 1) if total_classes else 0.0
        rows.append({
            "usn": usn,
            "name": data.get("name", ""),
            "total": total_classes,
            "attended": attended,
            "percentage": pct,
            "status": "Shortage" if pct < 75 else "Regular",
        })
    return rows


def _marks_rows(subject_code: str) -> list[dict]:
    rows = []
    for doc in db.collection(STUDENTS).order_by("usn").stream():
        student = doc.to_dict() or {}
        usn = (student.get("usn") or doc.id).upper()
        mark_doc = db.collection(MARKS).document(f"{usn}_{subject_code}").get()
        marks = mark_doc.to_dict() if mark_doc.exists else {}
        cie1 = marks.get("cie1") or 0
        cie2 = marks.get("cie2") or 0
        assignment = marks.get("assignment") or 0
        see = marks.get("see")
        avg = round((cie1 + cie2) / 2, 1)
        total_internal = round(avg + assignment, 1)
        total = round(total_internal + ((see or 0) / 2), 1) if see is not None else None
        grade_pct = total if total is not None else (total_internal / 50) * 100 if total_internal else 0
        rows.append({
            "usn": usn,
            "name": student.get("name", ""),
            "cie1": cie1,
            "cie2": cie2,
            "avg": avg,
            "assignment": assignment,
            "total_internal": total_internal,
            "see": "" if see is None else see,
            "total": "" if total is None else total,
            "grade": _grade_from_percentage(grade_pct),
        })
    return rows


def _pdf_response(filename: str, title: str, subtitle: str, headers: list[str], rows: list[list], shortage_col: int | None = None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Ramaiah Institute of Technology", styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(subtitle, styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ])
    if shortage_col is not None:
        for idx, row in enumerate(rows, start=1):
            if row[shortage_col] == "Shortage":
                table_style.add("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FEE2E2"))
    table.setStyle(table_style)
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}", styles["Normal"]))
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


def _excel_response(filename: str, sheet_name: str, headers: list[str], rows: list[list], shortage_col: int | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    for row in rows:
        ws.append(row)
        if shortage_col is not None and row[shortage_col] == "Shortage":
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(width, 28)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/attendance/pdf/<subject_code>", methods=["GET"])
def attendance_pdf(subject_code):
    subject_code = subject_code.upper()
    rows = _attendance_rows(subject_code)
    table_rows = [[r["usn"], r["name"], r["total"], r["attended"], f'{r["percentage"]}%', r["status"]] for r in rows]
    return _pdf_response(
        f"attendance_{subject_code}.pdf",
        f"Attendance Report - {_subject_name(subject_code)}",
        f"Subject Code: {subject_code}",
        ["USN", "Name", "Total Classes", "Attended", "Percentage", "Status"],
        table_rows,
        shortage_col=5,
    )


@export_bp.route("/attendance/excel/<subject_code>", methods=["GET"])
def attendance_excel(subject_code):
    subject_code = subject_code.upper()
    rows = _attendance_rows(subject_code)
    table_rows = [[r["usn"], r["name"], r["total"], r["attended"], r["percentage"], r["status"]] for r in rows]
    return _excel_response(
        f"attendance_{subject_code}.xlsx",
        "Attendance",
        ["USN", "Name", "Total Classes", "Attended", "Percentage", "Status"],
        table_rows,
        shortage_col=5,
    )


@export_bp.route("/marks/pdf/<subject_code>", methods=["GET"])
def marks_pdf(subject_code):
    subject_code = subject_code.upper()
    rows = _marks_rows(subject_code)
    table_rows = [[r["usn"], r["name"], r["cie1"], r["cie2"], r["avg"], r["assignment"], r["total_internal"], r["see"], r["total"], r["grade"]] for r in rows]
    return _pdf_response(
        f"marks_{subject_code}.pdf",
        f"Marks Report - {_subject_name(subject_code)}",
        f"Subject Code: {subject_code}",
        ["USN", "Name", "CIE1", "CIE2", "Avg", "Assignment", "Total Internal", "SEE", "Total", "Grade"],
        table_rows,
    )


@export_bp.route("/marks/excel/<subject_code>", methods=["GET"])
def marks_excel(subject_code):
    subject_code = subject_code.upper()
    rows = _marks_rows(subject_code)
    table_rows = [[r["usn"], r["name"], r["cie1"], r["cie2"], r["avg"], r["assignment"], r["total_internal"], r["see"], r["total"], r["grade"]] for r in rows]
    return _excel_response(
        f"marks_{subject_code}.xlsx",
        "Marks",
        ["USN", "Name", "CIE1", "CIE2", "Avg", "Assignment", "Total Internal", "SEE", "Total", "Grade"],
        table_rows,
    )
